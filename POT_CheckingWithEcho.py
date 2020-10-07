from selenium import webdriver
from time import sleep, time
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Color, PatternFill
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
import re

import functools

file_name = r"C:\Users\willlee\Downloads\LIST POT 2020.xlsx"
PATH = "C:\Program Files (x86)\chromedriver.exe"


class Excel_Operation:

    def add_in_color(self, column_to_write, row_to_start, color):
        write_position = column_to_write + row_to_start
        if color == 'Green':
            self.work_sheet[write_position].fill = self.greenfile
        elif color == 'Yellow':
            self.work_sheet[write_position].fill = self.yellowfile
        else:
            self.work_sheet[write_position].fill = self.redfile

    def replace_alphabet(self, text):
        return re.sub(r'[a-zA-Z]', r'*', text)

    def __init__(self, file_name, sheet_name):
        self.file_name = file_name
        self.sheet_name = sheet_name
        self.greenfile = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        self.yellowfile = PatternFill(start_color='008000', end_color='008000', fill_type='solid')
        self.redfile = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

    def load_workbook_with_sheet_name(self):
        self.book = openpyxl.load_workbook(self.file_name)
        self.work_sheet = self.book[self.sheet_name]

    def cell_to_read(self, column_to_read, row_to_start):
        # This function is assumed both read and write having the same row thus just have one row_to_start variable
        read_position = column_to_read + row_to_start  # C2
        try:
            text = self.work_sheet[read_position].value
            # expected new string ABCDEF*-01*
            new_string = self.replace_alphabet(text)
            # replace last character back
            print(new_string)
            return new_string, text
        except:
            return "", ""  # return empty string

    def cell_to_write(self, column_to_write, row_to_start, value_to_write_in):
        write_position = column_to_write + row_to_start  # D2
        self.work_sheet[write_position].value = value_to_write_in

    def save_file(self, file_name):
        self.book.save(file_name)


class webpage_Extract:
    # load the library from website
    def webpage_refresh(self, driver):
        sleep(1)
        # memory error for firefox
        page_source = driver.page_source
        sleep(1)
        soup = BeautifulSoup(page_source, 'lxml')
        self.soup = soup
        self.driver = driver
        return soup

    def check_all(self, *column_to_check):
        twoD_array = []
        oneD_array = []
        print(column_to_check)
        # convert to list
        list_column_to_check = list(column_to_check)
        items = self.soup.find_all('div', id='contenttablejqxGrid')
        for item in range(len(items[0].contents)):  # <--- This is the 25 elements
            print(items[0].contents[item])  # <--- This is 1 of the 25 element
            # go into contents
            # check if  text is empty if yes, break the loop
            if items[0].contents[item].contents[0].text == "":
                break
            else:
                for individual in list_column_to_check:
                    print(items[0].contents[item].contents[individual].text)
                    oneD_array.append(items[0].contents[item].contents[individual].text)
                twoD_array.append(oneD_array)
            # reset back 1D array
            oneD_array = []
        return twoD_array

    def convert_list_to_dict(self, twoD_list):
        # 1. look for unique value in 2D list
        list_to_dict = {}
        unique_value = list(set([i[1] for i in twoD_list]))
        # create key for the dictionary and empty list
        for element in unique_value:
            list_element = [i[0] for i in twoD_list if i[1] == element]
            list_to_dict[element] = list_element
        return list_to_dict


class Control_Echo:

    def __init__(self):
        self.driver = webdriver.Chrome(PATH)
        print('init')

    def launch_address(self, address):
        self.driver.get(address)
        print("open " + address)

    def webpage_control(self, id_name, action, value=""):

        var = self.driver.find_element_by_id(id_name)
        if action == "click":
            var.click()
            return True

        if action == "write":
            var.send_keys(value)
            return True

        else:
            print("match nothing")
            return False

    def buffer_time(self, timeOut_value):
        w = WebDriverWait(self.driver, timeOut_value)
        w.until(ec.presence_of_element_located((By.ID, "contenttablejqxGrid")))

    def wait(self, wait_time):
        self.driver.implicitly_wait(wait_time)

    @property
    def get_driver(self):
        print("get _driver ")
        return self.driver

    def teardown(self):
        self.driver.close()
        self.driver.quit()
        print("Finished")


# put an decorator for time recording
def logging_time(func):
    @functools.wraps(func)
    def wrapper_time(*args):
        start_time = time()
        func(*args)
        end_time = time()
        run_time = end_time - start_time
        print("--- %s seconds ---" % (time() - start_time))

    return wrapper_time


def look_for_any_lifecycle(dictionary, life_cycle):
    return dictionary[life_cycle]


@logging_time
def main():
    list_life_cycle = ['Active', 'Standard Support', 'LTB', 'Preliminary', 'Discontinued']
    successfully = False
    # ----------Webpage Operation-------------------------------
    webpage_ext = webpage_Extract()
    # ---------Excel Operation---------------------------------
    column_to_read = "C"
    column_to_write = "D"
    row_to_start = "2"
    row_to_end = "300"
    excel_operation = Excel_Operation(file_name, r"LIST POT 2020")
    excel_operation.load_workbook_with_sheet_name()
    # ---------Webpage Operation-------------------------------
    echo = Control_Echo()  # init webdriver
    # ---------------------------------------------------------
    echo.launch_address('https://echo.natinst.com/')

    print('opened Echo')
    echo.wait(5)
    successfully = echo.webpage_control(id_name='i0116', action='write',
                                        value='william.lee@ni.com')
    if successfully:
        print("Email Id Entered")
        successfully = False
    echo.wait(5)

    successfully = echo.webpage_control(id_name='idSIButton9', action='click')
    if successfully:
        print("Button Pressed")
        successfully = False
    echo.wait(5)

    print("User need to enter email and password id at the message box due to no way to control")
    input("Once Enter , press any key to continue")
    echo.wait(5)

    while True:
        # for i in range(row_to_start,row_to_end,1):
        print(row_to_start)
        wild_char_text, text = excel_operation.cell_to_read(column_to_read, row_to_start)
        if text == "":
            break
        else:

            successfully = echo.webpage_control(id_name='inputPart', action='write',
                                                value=wild_char_text)
            if successfully:
                print("Part Number Entered")
                successfully = False
            echo.wait(5)

            successfully = echo.webpage_control(id_name='lifecycleCheckbox972', action='click')
            if successfully:
                print("Discontinued Checked")
                successfully = False
            echo.wait(3)

            successfully = echo.webpage_control(id_name='searchButton', action='click')
            if successfully:
                print("Search button Clicked")
                successfully = False

            echo.buffer_time(3)

            # driver = echo.get_driver()
            soup = webpage_ext.webpage_refresh(echo.driver)
            two_d_list = webpage_ext.check_all(0, 5)
            # from table whcih one is active
            # take out all active put inside dictionary
            print(two_d_list)
            unqiue_dictionary_value = webpage_ext.convert_list_to_dict(two_d_list)
            print(unqiue_dictionary_value)
            unique_key = list(set(unqiue_dictionary_value))
            # Rule
            # 1. look the list and return a list
            # 2. from list look for the element == text
            for life_cycle in unique_key:
                if text in unqiue_dictionary_value[life_cycle]:
                    if life_cycle == 'Active':
                        # if active will not show other Active
                        excel_operation.cell_to_write(column_to_write, row_to_start, "Active")
                        excel_operation.add_in_color(column_to_write, row_to_start, 'Green')
                        break
                    elif life_cycle == 'Standard Support':
                        statement = 'Currently at Standard Support. '
                    elif life_cycle == 'LTB':
                        statement = 'Currently at LTB. '
                    elif life_cycle == 'Final Production':
                        statement = 'Currently at Final Production. '
                    elif life_cycle == 'Preliminary':
                        statement = 'Currently at Preliminary. '
                    elif life_cycle == 'Discontinued':
                        statement = 'Currently at Discontinued. '
                    else:
                        statement = 'No Match.....  check yourself.'
                    # get back active lifecycle
                    try:
                        # suggestion part number with active only #lost the flexibility of suggesting other part number
                        # basically dictionary should have everything just see how to get
                        # set rules which one come top (should be get from dictionary Active --> 'Standard Support' --> 'Final Production' --> 'LTB' --> 'Preliminary'
                        unique_key = list(set(unqiue_dictionary_value))
                        if "Active" in unique_key:
                            value = 'Active'
                        elif "Standard Support" in unique_key:
                            value = 'Standard Support'
                        elif "Final Production" in unique_key:
                            value = 'Final Production'
                        elif "LTB" in unique_key:
                            value = 'LTB'
                        elif "Preliminary" in unique_key:
                            value = 'Preliminary'
                        else:
                            value = 'Not Found any latest'
                        statement = statement + value + ' part number is :' + str(unqiue_dictionary_value[value])
                    except KeyError:
                        pass
                    excel_operation.cell_to_write(column_to_write, row_to_start, statement)
                    break
                else:
                    pass
                    # statement = 'Your part number no Match... Please check yourself.'
                    # #red line
                    # excel_operation.cell_to_write(column_to_write, row_to_start, statement)
                    # excel_operation.add_in_color(column_to_write, row_to_start, 'Red')

            # for key, values in unqiue_dictionary_value.items():
            #     # value should be a list
            #     # look for amount of the active
            #     for value in values:
            #         if value == text:  # if found the value , don't care in any of lifecycle
            #             # take out key value
            #             if key == 'Active':
            #                 if len(values) > 1:
            #                     # if more than one values then it should be check the latest one
            #                     pass
            #                 else:
            #                     statement = 'Active'
            #
            #                     break
            #
            #
            #  if the key value is active then having 1 only do something
            #             # else will get the latest one or all
            #             elif key == 'Standard Support':
            #                 statement = 'Currently at Standard Support'
            #                 # get back active lifecycle
            #                 look_for_any_lifecycle(dictionary, life_cycle)
            #             elif key == 'LTB':
            #                 statement = 'Currently at LTB'
            #             elif key == 'Final Production':
            #                 statement = 'Currently at Final Production'
            #             elif key == 'Preliminary':
            #                 statement = 'Currently at Preliminary'
            #             elif key == 'Discontinued':
            #                 statement = 'Currently at  Preliminary'
            #             else:
            #                 statement = 'No Match... Please check yourself.'
            #         else:
            #             pass
            #
            # # do what? if it look for most active Check if active same with the text need to take part number to check whether it is active
            # for i in two_d_list:
            #     if i[0] == text:  # if current Part number is in Active Life Cycle
            #         # take out the active one and do comparison with the excel if same , then write active if not same then write latest one
            #         if i[1] == "Active":
            #             excel_operation.cell_to_write(column_to_write, row_to_start, "Active")
            #             excel_operation.add_in_color(column_to_write, row_to_start)
            #             break
            #         else:  # able to search but not in active list
            #             # show it as what status
            #             excel_operation.cell_to_write(column_to_write, row_to_start, i[1])
            #     else:  # this is not active one
            #         # look for the active
            #         if i[1] == "Active" or i[1] == "Standard Support" or i[1] == "LTB" or i[1] == "Final Production":
            #             if i[1] == "Active":
            #                 statement = "Active part number is " + i[0]
            #             elif i[1] == "Standard Support":
            #                 statement = "Standard Support part number is " + i[0]
            #             elif i[1] == "LTB":
            #                 statement = "LTB part number is " + i[0]
            #             elif i[1] == "Final Production":
            #                 statement = "Final Production part number is " + i[0]
            #             elif i[1] == "Preliminary":
            #                 statement = "Preliminary part number is " + i[0]
            #             else:
            #                 statement = ""
            #            excel_operation.cell_to_write(column_to_write, row_to_start, statement)

            successfully = echo.webpage_control(id_name='clearSearchButton', action='click')
            if successfully:
                print("Clear button Clicked")
                row_to_start = int(row_to_start) + 1
                row_to_start = str(row_to_start)
                excel_operation.save_file(r"C:\Users\willlee\Downloads\LIST POT 2020_latest.xlsx")  # Save content
                successfully = False
    echo.teardown()


if __name__ == "__main__":
    main()
